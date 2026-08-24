#!/usr/bin/env python3
"""Expand Google Maps links in an Excel workbook and extract coordinates."""

import argparse
import logging
import re
import sys
from pathlib import Path
from tkinter import Tk, filedialog

import pandas as pd
import requests
from openpyxl import load_workbook
from tqdm import tqdm

OUTPUT_FILE = "mapmylink.xlsx"
OUTPUT_SHEET = "Map Coordinates"
LOG_FILE = "mapmylink.log"
REQUEST_TIMEOUT = 20
REQUIRED_COLUMNS = [
    "Head of Family",
    "Contact Address",
    "Prayer Group",
    "Maps Link",
]


def setup_logging(log_file: str = LOG_FILE) -> None:
    """Configure logging to both a file and stdout."""
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s | %(levelname)s | %(message)s",
        handlers=[
            logging.FileHandler(log_file, encoding="utf-8"),
            logging.StreamHandler(sys.stdout),
        ],
    )


def select_excel_file() -> str:
    """Open a graphical file picker and return the selected workbook path."""
    try:
        root = Tk()
        root.withdraw()
        root.attributes("-topmost", True)
        file_path = filedialog.askopenfilename(
            title="Select Input Excel File",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        )
        root.destroy()
        return file_path
    except Exception as exc:  # pragma: no cover - GUI failure is environment-specific
        logging.error("Could not open file picker: %s", exc)
        return ""


def is_valid_maps_link(url) -> bool:
    """Return True for non-empty HTTP(S) URLs."""
    if url is None:
        return False
    value = str(url).strip()
    if not value or value.lower() in {"nan", "none", "null"}:
        return False
    return value.startswith("http://") or value.startswith("https://")


def expand_google_maps_url(url: str, session: requests.Session):
    """Follow redirects and return (expanded_url, error_message)."""
    if not is_valid_maps_link(url):
        return "", "Empty or malformed URL"

    try:
        response = session.get(
            str(url).strip(),
            allow_redirects=True,
            timeout=REQUEST_TIMEOUT,
        )
        response.raise_for_status()
        if not response.url:
            return "", "No final URL returned"
        return response.url, ""
    except requests.exceptions.Timeout:
        return "", "Request timed out"
    except requests.exceptions.TooManyRedirects:
        return "", "Too many redirects"
    except requests.exceptions.RequestException as exc:
        return "", f"Request failed: {exc}"
    except Exception as exc:
        return "", f"Unexpected error: {exc}"


def extract_coordinates(url: str):
    """Extract latitude, longitude, and source from a Google Maps URL.

    Priority:
      1. !3dLAT!4dLON  - actual place coordinates
      2. q=LAT,LON      - query coordinates
      3. ll=LAT,LON     - ll coordinates
      4. @LAT,LON       - viewport coordinates (fallback only)
    """
    if not url:
        return "", "", ""

    matches = re.findall(
        r"!3d(-?\d+(?:\.\d+)?)!4d(-?\d+(?:\.\d+)?)",
        url,
        flags=re.IGNORECASE,
    )
    if matches:
        latitude, longitude = matches[-1]
        return latitude, longitude, "place"

    match = re.search(
        r"[?&]q=(-?\d+(?:\.\d+)?),(-?\d+(?:\.\d+)?)",
        url,
        flags=re.IGNORECASE,
    )
    if match:
        return match.group(1), match.group(2), "query"

    match = re.search(
        r"[?&]ll=(-?\d+(?:\.\d+)?),(-?\d+(?:\.\d+)?)",
        url,
        flags=re.IGNORECASE,
    )
    if match:
        return match.group(1), match.group(2), "ll"

    match = re.search(
        r"@(-?\d+(?:\.\d+)?),(-?\d+(?:\.\d+)?)",
        url,
        flags=re.IGNORECASE,
    )
    if match:
        return match.group(1), match.group(2), "viewport"

    return "", "", ""


def validate_coordinates(latitude, longitude) -> bool:
    """Validate geographic coordinate ranges."""
    try:
        lat = float(latitude)
        lon = float(longitude)
    except (TypeError, ValueError):
        return False
    return -90 <= lat <= 90 and -180 <= lon <= 180


def decimal_to_dms(value, is_latitude: bool = True) -> str:
    """Convert decimal degrees to Google Maps-style DMS notation."""
    if value is None or str(value).strip() == "":
        return ""
    try:
        value = float(value)
    except (TypeError, ValueError):
        return ""

    direction = (
        ("N" if value >= 0 else "S")
        if is_latitude
        else ("E" if value >= 0 else "W")
    )

    value = abs(value)
    degrees = int(value)
    minutes_float = (value - degrees) * 60
    minutes = int(minutes_float)
    seconds = round((minutes_float - minutes) * 60, 1)

    if seconds >= 60:
        seconds = 0
        minutes += 1
    if minutes >= 60:
        minutes = 0
        degrees += 1

    return f'{degrees}°{minutes:02d}\'{seconds:04.1f}"{direction}'


def process_excel(input_file: str, output_file: str | None = None):
    """Process the workbook and add/update the Map Coordinates sheet."""
    input_path = Path(input_file)
    if not input_path.exists():
        raise FileNotFoundError(f"Input file does not exist: {input_file}")

    output_path = Path(output_file) if output_file else input_path.parent / OUTPUT_FILE

    logging.info("Reading input file: %s", input_path)
    df = pd.read_excel(input_path, sheet_name=0)

    missing_columns = [column for column in REQUIRED_COLUMNS if column not in df.columns]
    if missing_columns:
        raise ValueError("Missing required columns: " + ", ".join(missing_columns))

    logging.info("Rows found: %d", len(df))

    session = requests.Session()
    session.headers.update(
        {
            "User-Agent": (
                "Mozilla/5.0 (X11; Linux x86_64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/120 Safari/537.36"
            )
        }
    )

    results = []
    successful = failed = empty = 0

    for index, row in tqdm(
        df.iterrows(), total=len(df), desc="Processing Maps Links", unit="link"
    ):
        excel_row = index + 2
        original_link = "" if pd.isna(row["Maps Link"]) else str(row["Maps Link"]).strip()

        base_result = {
            "Head of Family": row["Head of Family"],
            "Contact Address": row["Contact Address"],
            "Prayer Group": row["Prayer Group"],
            "Maps Link": original_link,
            "Expanded Maps Link": "",
            "Decimal Latitude": "",
            "Decimal Longitude": "",
            "DMS Latitude": "",
            "DMS Longitude": "",
        }

        if not original_link:
            logging.warning("Row %d: Maps Link is empty", excel_row)
            empty += 1
            results.append(base_result)
            continue

        expanded_url, expansion_error = expand_google_maps_url(original_link, session)
        if expansion_error:
            logging.error(
                "Row %d: Could not expand URL: %s | %s",
                excel_row,
                original_link,
                expansion_error,
            )
            failed += 1
            results.append(base_result)
            continue

        base_result["Expanded Maps Link"] = expanded_url
        latitude, longitude, source = extract_coordinates(expanded_url)

        if not latitude or not longitude:
            logging.error("Row %d: Coordinates not found: %s", excel_row, expanded_url)
            failed += 1
            results.append(base_result)
            continue

        if not validate_coordinates(latitude, longitude):
            logging.error(
                "Row %d: Invalid coordinates: %s, %s",
                excel_row,
                latitude,
                longitude,
            )
            failed += 1
            results.append(base_result)
            continue

        if source == "viewport":
            logging.warning(
                "Row %d: Only viewport coordinates were available; verify manually.",
                excel_row,
            )

        base_result["Decimal Latitude"] = float(latitude)
        base_result["Decimal Longitude"] = float(longitude)
        base_result["DMS Latitude"] = decimal_to_dms(latitude, True)
        base_result["DMS Longitude"] = decimal_to_dms(longitude, False)

        logging.info(
            "Row %d: %s, %s | source=%s",
            excel_row,
            latitude,
            longitude,
            source,
        )
        successful += 1
        results.append(base_result)

    output_df = pd.DataFrame(results)

    workbook = load_workbook(input_path)
    if OUTPUT_SHEET in workbook.sheetnames:
        del workbook[OUTPUT_SHEET]
    worksheet = workbook.create_sheet(OUTPUT_SHEET)

    for column_number, column_name in enumerate(output_df.columns, start=1):
        worksheet.cell(row=1, column=column_number).value = column_name

    for row_number, row_data in enumerate(output_df.itertuples(index=False), start=2):
        for column_number, value in enumerate(row_data, start=1):
            worksheet.cell(
                row=row_number,
                column=column_number,
            ).value = "" if pd.isna(value) else value

    for column in worksheet.columns:
        max_length = max(
            (len(str(cell.value)) for cell in column if cell.value is not None),
            default=0,
        )
        worksheet.column_dimensions[column[0].column_letter].width = min(max_length + 2, 60)

    worksheet.freeze_panes = "A2"
    workbook.save(output_path)

    logging.info("=" * 60)
    logging.info("PROCESS COMPLETED")
    logging.info("Total rows: %d", len(df))
    logging.info("Successful: %d", successful)
    logging.info("Failed: %d", failed)
    logging.info("Empty links: %d", empty)
    logging.info("Output: %s", output_path)
    logging.info("Log file: %s", Path(LOG_FILE).absolute())
    logging.info("=" * 60)

    return output_path


def parse_arguments():
    parser = argparse.ArgumentParser(
        description="Expand Google Maps links from an Excel file and extract coordinates."
    )
    parser.add_argument(
        "input_file",
        nargs="?",
        help="Input Excel file. If omitted, a file picker is opened.",
    )
    parser.add_argument(
        "-o",
        "--output",
        help="Optional output Excel filename/path. Defaults to mapmylink.xlsx next to input.",
    )
    return parser.parse_args()


def main() -> None:
    setup_logging()
    args = parse_arguments()
    input_file = args.input_file or select_excel_file()

    if not input_file:
        logging.info("No input file selected.")
        return

    try:
        process_excel(input_file, args.output)
    except Exception as exc:
        logging.exception("Fatal error: %s", exc)
        raise SystemExit(1) from exc


if __name__ == "__main__":
    main()
